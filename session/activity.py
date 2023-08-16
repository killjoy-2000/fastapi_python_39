import openpyxl

# try:
#     wb = openpyxl.load_workbook("test_xl_py.xlsx")
#     sheet = wb["Sheet2"]
#     def check_sts(ip):
#         # print(ip)
#         try:
#             search_value = ip
#             found_cells = []
#             col = sheet["A"]
#             # for row in sheet.iter_rows():
#             #     for cell in row:
#             #         # if cell != search_value:
#             #         #     s = "data inserted"
#             #         #     
#             #         #     sheet.append(put_data)
#             #         #     wb.save("log.xlsx")
#             #         #     return s
#             #         if cell.value == search_value:
#             #             s = "data found"
#             #             found_cells.append(row)
#             #             return s
#             for cell in col:
#                 if cell.value == search_value:
#                     if row[1].value == "T":
#                         return True
#                     # cell.value = None
#                     # sheet.delete_rows(cell.row, 1)
#                     # wb.save("test_xl_py.xlsx")
#                     # if row[1].value == "T":
#                     #     row[1].value = "F"
#                     # return True
#                 # else:
#             put_data = {
#                     "A": search_value,
#                     "B" : "T",
#                     "C" : "T",
#                 }
#             sheet.append(put_data)
#             wb.save("test_xl_py.xlsx")
#             return False
#         except Exception as e:
#             print(str(e))
#             return str(e)

#         # return "sec"
#     # print(sheet["A2"].value)

#     def reset_activity():
#         for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3):
#         # Check if the value in column A is not empty
#             if row[0].value is not None:
#             # Set the values in column B and C to F
#                 row[1].value = 'F'
#                 row[2].value = 'F'
#         wb.save("test_xl_py.xlsx")


# except Exception as e:
#     print(str(e))

# finally:
#     wb.close()


try:
    wb = openpyxl.load_workbook("test_xl_py.xlsx")
    sheet = wb["Sheet2"]

    def clear_activity():
        try:
            for row in sheet.iter_rows(min_row=1):  # Assuming headers are in the first row
                column_a_value = row[0].value
                if column_a_value:
                    row[1].value = 'F'  # Column B
                    row[2].value = 'F'  # Column C
            return "data activity formatted"
        except Exception as e:
            print(str(e))
            return str(e)
        finally:
            wb.save("test_xl_py.xlsx")
            wb.close()
    
    def check_sts(ip):
        login_sts = "F"
        try:
            search_val = ip
            found_cells = []
            for row in sheet.iter_rows(min_row=1):
                column_a_val = row[0].value
                column_b_val = row[1].value
                
                if column_a_val == search_val:
                    login_sts = column_b_val
                    # break
                    if(login_sts == "T"):
                        return "available"
                    else:
                        # convert = list(row)
                        # row[1] = "T"
                        # convert = tuple(convert)
                        row[1].value = 'T'
                        wb.save("test_xl_py.xlsx")
                        wb.close()
                        return "not available"                        

        except Exception as e:
            print(str(e))
            return str(e)



except Exception as e:
    print(str(e))

finally:
    wb.close()