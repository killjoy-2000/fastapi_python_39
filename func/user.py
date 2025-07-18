# import pandas as pd

# df = pd.read_excel("test_xl_py.xlsx")
# print(df)


import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("test_xl_py.xlsx")
# print(wb.sheetnames)
# sheet = wb["Sheet1"]
sheet = wb["Sheet1"]
print(sheet.title)
print(sheet["A1"].value)

def add_user(data):
    try:
        search_val = data.user_id
        found_cells = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
            column_b_val = row[1].value
            if column_b_val == search_val:
                return "user_exist"
            # elif column_a_val != search_val:
        data_to_put = {
            "A" : data.user,
            "B" : data.user_id,
            "C" : data.image_no,
            "D" : data.image_consumed,
            "E" : data.tag,
        }
        sheet.append(data_to_put)
        wb.save("test_xl_py.xlsx")
        return True
    except Exception as e:
        print(str(e))
        return "error >>>>>>>>>> " + str(e)

    # finally:
    #     wb.close()

# for row in sheet.iter_rows(values_only=True):
#     for cell in row:
#         print(cell)

# data_to_put = {
#     "A" : "Rudra",
#     "B" : 105,
#     "C" : 200,
#     "D" : 34,
#     "E" : "F",
# }

# search_value = "Anwesa"
# found_cells = []
# for row in sheet.iter_rows(values_only=True):
#     for cell in row:
#         if cell == search_value:
#             found_cells.append(row)

#             # cell_number = f"{get_column_letter(cell.column)}{cell.row}"
#             # print(cell_number)

# # Print the found cells
# for row in found_cells:
#     print(row)
#     print(row[3])
# # for cell in found_cells:
# #     print(cell)
# #     print(found_cells)

# # sheet.append(data_to_put)
# # wb.save("test_xl_py.xlsx")


# wb.close()
